import string
import openpyxl
import os
import shutil


workbook = openpyxl.load_workbook('./liste-globale-des-PFE.xlsx')


def remove_non_english_chars(text):
    printable = set(string.printable) - set('&')
    new_string = ""
    for char in text:
        if char in printable and ord(char) < 128:
            new_string += char
    return new_string


def create_dir(year, number, number2, files_not_found):
    file_number = 1
    handle = '123456789/'
    collection = '123456789/2'

    print('Creating directory for year: ', year)

    worksheet = workbook['PFE '+str(year)]

    column_indices = {}

    for column in worksheet.iter_cols(min_row=1, max_row=1, values_only=True):
        for index, cell_value in enumerate(column):
            column_indices[cell_value] = index

    records = []
    records_meta = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        record = {}
        record_meta = {}
        for column_name, column_index in column_indices.items():
            if column_name == 'Titre':
                record['title'] = row[0]
            elif column_name == 'Auteur':
                record['author'] = row[1]
            elif column_name == 'Sujet':
                record['description'] = row[2]
            elif column_name == 'Date':
                record['date'] = row[3]
            elif column_name == 'Organisme':
                record_meta['organisme'] = row[4]
            elif column_name == 'Encadrants':

                record_meta['encadrants'] = row[5]

        records.append(record)
        records_meta.append(record_meta)
    for record in records:

        xml_string = '<?xml version="1.0" encoding="utf-8" standalone="no"?>\n'
        xml_string += '<dublin_core schema="dc">\n'
        title = record['title']
        link = None
        try:

            link = worksheet.cell(row=file_number+1, column=1).hyperlink.target
        except:
            print('No link found')
            files_not_found += f'LINK {title} for {year} not found \n \n'

        xml_string += f'<dcvalue element="identifier" qualifier="uri">https:&#x2F;&#x2F;sbn.inpt.ac.ma&#x2F;handle&#x2F;123456789&#x2F;{number}</dcvalue>\n'
        for key, value in record.items():
            if key == "author":
                if value is not None:
                    value = remove_non_english_chars(str(value))
                xml_string += f'<dcvalue element="contributor" qualifier="author">{value}</dcvalue>\n'
            else:
                if value is not None:
                    value = remove_non_english_chars(str(value))
                xml_string += f'<dcvalue element="{key}" qualifier="none">{value}</dcvalue>\n'
        xml_string += '</dublin_core>'

        file = ''
        if link is not None:
            link = link.replace('../../Sauvegarde/', '')
            link_import = link.replace('PFE '+str(year)+'/', '')
            link_copy = remove_non_english_chars(link_import)
            text_import = f'{link_copy}	bundle:ORIGINAL'

            try:
                os.makedirs(f'./csv/{year}/{number}', exist_ok=True)
                shutil.copy(f'./PFE/{link}',
                            f'./csv/{year}/{number}/{link_copy}')
                with open(f'./csv/{year}/{number}/contents', 'w', encoding='utf-8') as f:
                    f.write(text_import)
                with open(f'./csv/{year}/{number}/dublin_core.xml', 'w', encoding='utf-8') as f:
                    f.write(xml_string)
                with open(f'./csv/{year}/{number}/collections', 'w', encoding='utf-8') as f:
                    f.write(collection)
                with open(f'./csv/{year}/{number}/handle', 'w', encoding='utf-8') as f:
                    f.write(handle+str(number))

            except:
                print('File not found')
                files_not_found += f'FILE {title} for {year} not found \n \n'
            file_number += 1
            number += 1
    for record in records_meta:
        xml_string = '<dublin_core schema="pfe">\n'
        for key, value in record.items():
            if value is not None:
                value = remove_non_english_chars(str(value))
                value = value.replace('&', 'and')
            if key == 'encadrants':
                if value is None:
                    value = ''
                if value is not None:
                    value = remove_non_english_chars(str(value))
                    value = value.replace('&', 'and')
                name_list = value.split('\n')
                for name in name_list:
                    text = name.strip()
                    xml_string += f'<dcvalue element="{key}" qualifier="none">{text}</dcvalue>\n'
            else:
                xml_string += f'<dcvalue element="{key}" qualifier="none">{value}</dcvalue>\n'
        xml_string += '</dublin_core>'
        try:
            with open(f'./csv/{year}/{number2}/metadata_pfe.xml', 'w', encoding='utf-8') as f:
                f.write(xml_string)
        except:
            print('File not found')
        number2 += 1
    return number, number2, files_not_found


number = 12000
number2 = 12000
files_not_found = ''
# years from 1999 to 2020
years = [1999, 2000, 2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008,
         2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019]
for year in years:
    number, number2, files_not_found = create_dir(
        year, number, number2, files_not_found)
    print('Done for year: ', year)
    print('----------------------------------------')

with open(f'./csv/files_not_found.txt', 'w', encoding='utf-8') as f:
    f.write(files_not_found)
